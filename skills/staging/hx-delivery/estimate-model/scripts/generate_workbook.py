#!/usr/bin/env python3
"""
Generate Estimate Workbook

Creates a summary Excel workbook from sizing.json using openpyxl.
Produces a colour-coded model comparison table with aggregate metrics.

Usage:
    python ci/generate_workbook.py [base-path]

Defaults to "files/" relative to the repository root.
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Constants ----------------------------------------------------------------

_SIZE_LABEL = {"EXTRA_LARGE": "XL", "LARGE": "L", "MEDIUM": "M", "SMALL": "S"}
_SIZE_ORDER = {"EXTRA_LARGE": 4, "LARGE": 3, "MEDIUM": 2, "SMALL": 1}

_SIZE_COLORS = {
    "S": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "M": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "L": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "XL": PatternFill(start_color="D9C2EC", end_color="D9C2EC", fill_type="solid"),
}

_HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_BOLD_FONT = Font(bold=True)

_COLUMNS = [
    ("Model Name", 40),
    ("Size", 8),
    ("Score", 8),
    ("Confidence", 12),
    ("Formulas", 12),
    ("Sheets", 8),
    ("Max Depth", 10),
    ("VBA", 6),
    ("VBA Lines", 10),
    ("Integrations", 13),
    ("Flags", 50),
]


# --- Helpers ------------------------------------------------------------------

def _strip_extension(file_name: str) -> str:
    return file_name.rsplit(".", 1)[0] if "." in file_name else file_name


def _sort_models(models: list) -> list:
    return sorted(
        models,
        key=lambda m: (
            -_SIZE_ORDER.get(m.get("size", "SMALL"), 0),
            -(m.get("score") or 0),
            m.get("file_name", ""),
        ),
    )


def _apply_header_style(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER


def _apply_size_formatting(ws, col_letter: str, first_row: int, last_row: int):
    if first_row > last_row:
        return
    cell_range = f"{col_letter}{first_row}:{col_letter}{last_row}"
    for label, fill in _SIZE_COLORS.items():
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=[f'"{label}"'], fill=fill),
        )


def _write_model_table(ws, models: list) -> int:
    headers = [name for name, _ in _COLUMNS]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, model in enumerate(models, 2):
        size_raw = model.get("size", "SMALL")
        size_label = _SIZE_LABEL.get(size_raw, size_raw)

        values = [
            _strip_extension(model.get("file_name", "")),
            size_label,
            model.get("score", 0),
            model.get("confidence", "LOW"),
            model.get("formula_count", 0),
            model.get("sheet_count", 0),
            model.get("max_dependency_depth", 0),
            "Yes" if model.get("has_vba") else "No",
            model.get("vba_lines", 0),
            "Yes" if model.get("has_integrations") else "No",
            ", ".join(model.get("flags") or []),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = _THIN_BORDER

            # Alternating row shading
            if row_idx % 2 == 0:
                cell.fill = _ALT_ROW_FILL

        # Number formats
        ws.cell(row=row_idx, column=3).number_format = "0.0"
        ws.cell(row=row_idx, column=5).number_format = "#,##0"
        ws.cell(row=row_idx, column=9).number_format = "#,##0"

        # Text wrap on flags column
        ws.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True)

    return len(models) + 2  # next empty row


def _write_aggregates(ws, sizing_data: dict, start_row: int):
    row = start_row + 1  # blank row gap

    # Size distribution header
    ws.cell(row=row, column=1, value="Size Distribution").font = _BOLD_FONT
    row += 1

    dist = sizing_data.get("size_distribution", {})
    for size_key in ["SMALL", "MEDIUM", "LARGE", "EXTRA_LARGE"]:
        label = _SIZE_LABEL.get(size_key, size_key)
        count = dist.get(size_key, 0)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1  # blank row gap

    # Aggregate metrics header
    ws.cell(row=row, column=1, value="Aggregate Metrics").font = _BOLD_FONT
    row += 1

    metrics = [
        ("Total Formulas", sizing_data.get("total_formulas", 0)),
        ("Average Score", sizing_data.get("average_score", 0)),
        ("Total Models", len(sizing_data.get("models", []))),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if label == "Average Score":
            cell.number_format = "0.0"
        elif isinstance(value, int):
            cell.number_format = "#,##0"
        row += 1


# --- Public API ---------------------------------------------------------------

def generate_workbook(sizing_data: dict, output_dir: Path) -> Path:
    """
    Generate a summary Excel workbook from sizing data.

    Returns the Path to the created file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    models = _sort_models(sizing_data.get("models", []))

    # Write model table
    next_row = _write_model_table(ws, models)

    # Header styling
    _apply_header_style(ws, len(_COLUMNS))

    # Conditional formatting on size column (B)
    if models:
        _apply_size_formatting(ws, "B", 2, len(models) + 1)

    # Column widths
    for col_idx, (_, width) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Aggregates
    _write_aggregates(ws, sizing_data, next_row)

    # Save
    project_name = sizing_data.get("project_name", "Unknown")
    output_file = output_dir / f"Estimate Template - {project_name}.xlsx"
    wb.save(str(output_file))

    return output_file


# --- Main ---------------------------------------------------------------------

def main():
    _root = Path(__file__).parent.parent
    default_base = _root / "files"
    project_dir = Path(sys.argv[1]).resolve() if len(sys.argv) >= 2 else default_base.resolve()
    output_path = project_dir / "triage output"

    sizing_file = output_path / "sizing.json"
    if not sizing_file.exists():
        print(f"Error: {sizing_file} not found. Run send_to_claude.py first.", file=sys.stderr)
        sys.exit(1)

    try:
        with open(sizing_file, encoding="utf-8") as f:
            sizing_data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON in {sizing_file}: {e}", file=sys.stderr)
        sys.exit(1)

    print("Generating estimate workbook", file=sys.stderr)
    wb_path = generate_workbook(sizing_data, output_path)
    print(f"  {wb_path.name}", file=sys.stderr)
    print("  Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
