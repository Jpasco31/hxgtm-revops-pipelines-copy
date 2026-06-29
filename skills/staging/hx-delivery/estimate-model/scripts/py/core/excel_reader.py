"""
Excel Reader

Wraps openpyxl with the same interface as the JS ExcelReader:
- Cell iteration helpers
- Large file sampling (>20MB -> top 100 rows per sheet)
- Sheet visibility detection (visible / hidden / veryHidden)
- Named range enumeration
- VBA presence detection via ZIP inspection
- Error cell counting via raw XML scan (matches JS xlsx library behaviour)
"""

import os
import re
import zipfile
import warnings
from html import unescape as html_unescape
from pathlib import Path
from typing import Generator, Optional, Tuple

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.chartsheet import Chartsheet

LARGE_FILE_THRESHOLD_MB = 20
MAX_ROWS_FOR_LARGE_FILES = 100

# Excel error values as stored in XML / as cell values
EXCEL_ERRORS = frozenset({
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#GETTING_DATA"
})

# Regex to count error cells: <c ... t="e" ...> in worksheet XML
_ERROR_CELL_RE = re.compile(rb'<c\b[^>]*\bt="e"')

# Regex to count formula cells: any <f> or <f .../> element in worksheet XML.
# This counts both regular formulas and shared formula references (t="shared"),
# matching the JS xlsx library's cell.f behaviour.
_FORMULA_TAG_RE = re.compile(rb'<f(?=[> /\t\n\r])')


class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self._error_counts: dict[str, int] = {}   # sheet title -> error count from XML
        self.metadata = {
            "fileName": Path(file_path).name,
            "fileSizeMB": 0.0,
            "samplingUsed": False,
            "samplingMethod": "full-analysis",
        }

    def load(self) -> openpyxl.Workbook:
        try:
            size_bytes = os.path.getsize(self.file_path)
            self.metadata["fileSizeMB"] = round(size_bytes / (1024 * 1024), 2)

            is_large = self.metadata["fileSizeMB"] > LARGE_FILE_THRESHOLD_MB
            if is_large:
                self.metadata["samplingUsed"] = True
                self.metadata["samplingMethod"] = "top-100-rows"

            # data_only=False: formulas stored as strings (e.g. "=A1+B1")
            # keep_vba=False: don't try to parse VBA binary here
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                read_only=is_large,
                data_only=False,
                keep_vba=False,
            )

            # Single-pass XML scan: error cells + formula cells.
            # Bypasses openpyxl limitations with shared formula expansion and
            # error-type detection, matching JS xlsx library behaviour.
            self._scan_worksheet_xml()

            return self.workbook
        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file: {e}")

    def get_workbook(self) -> openpyxl.Workbook:
        if self.workbook is None:
            self.load()
        return self.workbook

    def get_metadata(self) -> dict:
        return self.metadata

    # -------------------------------------------------------------------------
    # Error cell detection via raw XML
    # -------------------------------------------------------------------------

    def _get_sheet_states_from_xml(self) -> dict:
        """
        Parse workbook.xml for sheet visibility states.
        openpyxl does not correctly read state='hidden' on Chartsheet objects,
        so we bypass it and read from the XML directly.
        Returns {sheet_name: state_string} where state is 'visible'/'hidden'/'veryHidden'.
        """
        try:
            with zipfile.ZipFile(self.file_path) as zf:
                data = zf.read("xl/workbook.xml")
            result = {}
            for m in re.finditer(rb'<sheet\b([^/]*)/>', data):
                attrs = m.group(1)
                name_m = re.search(rb'name="([^"]+)"', attrs)
                state_m = re.search(rb'state="([^"]+)"', attrs)
                if name_m:
                    # html_unescape handles &amp; &lt; &gt; &apos; &quot; in sheet names
                    name = html_unescape(name_m.group(1).decode("utf-8", errors="replace"))
                    state = state_m.group(1).decode("utf-8") if state_m else "visible"
                    result[name] = state
            return result
        except Exception:
            return {}

    def _scan_worksheet_xml(self) -> None:
        """
        Single-pass scan of all worksheet XMLs to collect:
        - Error cell count (t="e") — mirrors JS cell.t === 'e'
        - Formula cell count (<f> tags) — mirrors JS cell.f, including shared formulas
        Results stored in self._error_count and self._formula_count.
        """
        self._error_count = 0
        self._formula_count = 0
        try:
            with zipfile.ZipFile(self.file_path) as zf:
                sheet_files = sorted(
                    n for n in zf.namelist()
                    if re.match(r"xl/worksheets/sheet\d+\.xml$", n)
                )
                for sf in sheet_files:
                    data = zf.read(sf)
                    self._error_count += len(_ERROR_CELL_RE.findall(data))
                    self._formula_count += len(_FORMULA_TAG_RE.findall(data))
        except Exception:
            pass  # Non-fatal; counts will show as 0

    def _get_total_error_cells(self) -> int:
        return self._error_count

    def _get_total_formula_cells(self) -> int:
        return self._formula_count

    # -------------------------------------------------------------------------
    # Cell iteration
    # -------------------------------------------------------------------------

    def iter_cells(self, worksheet, max_rows: Optional[int] = None) -> Generator[Tuple[Cell, str], None, None]:
        """Yield (cell, address) for every non-empty cell (skips chart sheets)."""
        if isinstance(worksheet, Chartsheet):
            return
        effective_max = max_rows or (MAX_ROWS_FOR_LARGE_FILES if self.metadata["samplingUsed"] else None)
        for row in worksheet.iter_rows(max_row=effective_max):
            for cell in row:
                if cell.value is not None:
                    yield cell, cell.coordinate

    def iter_formula_cells(self, worksheet, max_rows: Optional[int] = None) -> Generator[Tuple[Cell, str, str], None, None]:
        """Yield (cell, formula_without_equals, address) for formula cells."""
        for cell, address in self.iter_cells(worksheet, max_rows):
            if isinstance(cell.value, str) and cell.value.startswith("="):
                yield cell, cell.value[1:], address  # strip leading =

    def iter_sheets(self) -> Generator[Tuple, None, None]:
        """Yield (worksheet, sheet_name) for all sheets (skips chart sheets)."""
        wb = self.get_workbook()
        for name in wb.sheetnames:
            ws = wb[name]
            if not isinstance(ws, Chartsheet):
                yield ws, name

    # -------------------------------------------------------------------------
    # Sheet visibility
    # -------------------------------------------------------------------------

    def get_sheet_visibility(self, sheet_name: str) -> str:
        """Returns 'visible', 'hidden', or 'veryHidden'."""
        wb = self.get_workbook()
        return wb[sheet_name].sheet_state

    def count_sheets_by_visibility(self) -> dict:
        wb = self.get_workbook()
        xml_states = self._get_sheet_states_from_xml()
        counts = {"total": 0, "visible": 0, "hidden": 0, "veryHidden": 0}
        for name in wb.sheetnames:
            counts["total"] += 1
            # Prefer XML-parsed state (more reliable for chart sheets)
            state = xml_states.get(name) or "visible"
            if state == "visible":
                counts["visible"] += 1
            elif state == "veryHidden":
                counts["veryHidden"] += 1
            else:
                counts["hidden"] += 1
        return counts

    # -------------------------------------------------------------------------
    # Named ranges
    # -------------------------------------------------------------------------

    def get_named_ranges(self) -> list:
        """
        Return ALL defined names by parsing workbook.xml directly.
        openpyxl's defined_names API deduplicates names that exist in multiple
        scopes (e.g. a sheet-scoped 'Bro' + a global 'Bro'), so we bypass it.
        Mirrors JS xlsx library's wb.Workbook.Names which counts all entries.
        """
        try:
            with zipfile.ZipFile(self.file_path) as zf:
                data = zf.read("xl/workbook.xml")
            matches = re.findall(rb'<definedName\b([^>]+)>', data)
            ranges = []
            for attrs in matches:
                name_m = re.search(rb'name="([^"]+)"', attrs)
                if name_m:
                    ranges.append({"name": name_m.group(1).decode("utf-8", errors="replace")})
            return ranges
        except Exception:
            # Fallback to openpyxl API if XML parse fails
            wb = self.get_workbook()
            return [{"name": defn.name} for defn in wb.defined_names.values()]

    # -------------------------------------------------------------------------
    # VBA detection (ZIP inspection - no parsing)
    # -------------------------------------------------------------------------

    def has_vba(self) -> bool:
        try:
            with zipfile.ZipFile(self.file_path) as z:
                return "xl/vbaProject.bin" in z.namelist()
        except Exception:
            return False

    # -------------------------------------------------------------------------
    # Statistics (mirrors JS getStatistics)
    # -------------------------------------------------------------------------

    def get_statistics(self) -> dict:
        visibility = self.count_sheets_by_visibility()
        named_ranges = self.get_named_ranges()

        # Count non-empty cells via openpyxl (excludes stub/formatting-only cells)
        non_empty_cells = 0
        for ws, _name in self.iter_sheets():
            for _cell, _address in self.iter_cells(ws):
                non_empty_cells += 1

        # Formula and error counts from XML — exact match to JS xlsx library.
        # openpyxl's shared formula expansion is incomplete; XML counting is authoritative.
        formula_cells = self._get_total_formula_cells()
        error_cells = self._get_total_error_cells()

        # totalCells = all non-empty cells (note: JS also counts empty stub cells
        # within the used range due to sheetStubs:true; Python's count is data-only)
        total_cells = non_empty_cells

        # inputCells = non-empty cells minus formula cells.
        # This correctly handles shared-formula cells that openpyxl mis-reads as values.
        input_cells = max(0, non_empty_cells - formula_cells)

        return {
            "fileName": self.metadata["fileName"],
            "fileSizeMB": self.metadata["fileSizeMB"],
            "samplingUsed": self.metadata["samplingUsed"],
            "samplingMethod": self.metadata["samplingMethod"],
            "sheetCount": visibility["total"],
            "visibleSheets": visibility["visible"],
            "hiddenSheets": visibility["hidden"],
            "veryHiddenSheets": visibility["veryHidden"],
            "totalCells": total_cells,
            "formulaCells": formula_cells,
            "inputCells": input_cells,
            "errorCells": error_cells,
            "namedRanges": len(named_ranges),
        }
